import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines

# 读取两个Excel文件
employee_info = pd.read_excel('员工基本信息表.xlsx')
employee_performance = pd.read_excel('员工绩效表.xlsx')

# 筛选2024年第4季度的绩效数据
performance_q4_2024 = employee_performance[
    (employee_performance['年度'] == 2024) & 
    (employee_performance['季度'] == 4)
]

# 合并两个表（使用left join保留所有员工信息）
merged_data = pd.merge(
    employee_info,
    performance_q4_2024[['员工ID', '绩效评分']],
    on='员工ID',
    how='left'
)

# 重命名绩效评分列
merged_data = merged_data.rename(columns={'绩效评分': '2024年第4季度绩效评分'})

# 将结果保存到新的Excel文件
output_file = '员工信息与绩效表.xlsx'
merged_data.to_excel(output_file, index=False)

# 加载工作簿以添加图表
wb = load_workbook(output_file)
ws = wb.active

# 创建条形图
chart = BarChart()
chart.title = "2024年第4季度员工绩效评分分布"
chart.y_axis.title = "绩效评分"
chart.x_axis.title = "员工姓名"

# 设置Y轴范围
chart.y_axis.scaling.min = 3.0  # 设置Y轴最小值
chart.y_axis.scaling.max = 5.0  # 设置Y轴最大值

# 添加网格线使数据更易读
chart.y_axis.majorGridlines = ChartLines()
chart.y_axis.majorUnit = 0.5  # 设置主要刻度间隔为0.5

# 设置数据范围（绩效评分）
data = Reference(ws, min_col=6, min_row=2, max_row=ws.max_row, max_col=6)
# 设置类别标签（员工姓名）
categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

# 添加数据到图表
chart.add_data(data)
chart.set_categories(categories)

# 设置图表样式
chart.height = 15  # 图表高度
chart.width = 25   # 图表宽度

# 将图表插入到工作表中
ws.add_chart(chart, "I2")

# 保存工作簿
wb.save(output_file)

print("已生成带有绩效图表的Excel文件！Y轴范围已调整为3.0-5.0")
print("\n合并后的数据预览：")
print(merged_data) 